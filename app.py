import streamlit as st
import pandas as pd
import numpy as np
from datetime import timedelta, datetime
import calendar
from io import BytesIO
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re

def date_to_week(date):
    dt = pd.to_datetime(date).to_pydatetime()
    weekday = dt.weekday()
    days_since_friday = (weekday - 4) % 7
    friday = dt - timedelta(days=days_since_friday)
    week_start = friday
    week_end = friday + timedelta(days=6)
    week_start_str = f"{week_start.day:02d} {calendar.month_name[week_start.month]}"
    week_end_str = f"{week_end.day:02d} {calendar.month_name[week_end.month]}"
    return f"{week_start_str} - {week_end_str}"
#insert data dictionary for functional location code
code_to_text = {
    "901":"Propulsion System",
    "902":"Control Circuit",
    "903":"Auxiliary Power Supply",
    "904":"ATP/AO Signalling",
    "905":"Lighting",
    "906":"Air Conditioning and Ventilation System",
    "907":"Onboard Communication System",
    "908":"Doors",
    "909":"Special Equipment",
    "911":"Carbody Shell",
    "912":"Carbody Interior",
    "913":"Carbody Exterior",
    "915":"Bogies",
    "916":"Brake System",
    "918":"Coupling and Interconnection",
    "922":"Train Management System(TMS)",
    "923":"Liquid Cooling System"}
# Function to get code text based on the code
def get_code_text(code):
    # if the code has less than 17 characters, return "Rolling Stock"
    if len(code) > 17:
    #if the 17th character to 20th character is in the dictionary, return the corresponding text
        code_text = code_to_text.get(code[16:19], "#N/A")
        return code_text
    else:
        return "Rolling Stock"

rules = {
    'Guide tire worn out': ['Guide tire', 'worn'],  # Must contain ALL words
    'Guide tire warning': ['Guide tire', 'warning'],
    'Load tire worn out':['load','worn'],
    'Load tire warning':['load','warning'],
    'CCD worn out': ['CCD', 'Worn out'],
    'Train not responding in TWP': ['TWP'],
    'Smoke alarm': ['Smoke alarm'],
    'APU faulty': ['APU faulty'],
    'Guide tire lost signal': ['Guide tire', 'signal'],
    'CCD replacement': ['Collector'],
    'CCD crack': ['CCD'],
    'Brake pressure low':['TWP','Brake'],
    'Ceiling loud noise':['ceiling','sound'],
    'Door major failure':['door major'],
    'Gangway lound noise':['gangway loud'],
    'Liquid cooling system':['liquid'],
    'MTC Major failure':['MTC','major','failure'],
    'Steering Cylinder leak':['steering','leak'],
    'Steering Cylinder warning':['steering','warning'],
    'Water dripping':['water','drip'],
    'Wheel Arc':['Allcar','arc'],
    'Wheel well door lock braket':['bracket']
    }
# Pre-compile regex patterns for speed
compiled_patterns = {}
for category, keywords in rules.items():
    # Regex: (?=.*word1)(?=.*word2)(?=.*word3)
    pattern = ''.join(f'(?=.*{keyword})' for keyword in keywords)
    compiled_patterns[category] = re.compile(pattern, flags=re.IGNORECASE)

# Apply classification
def classify_text(text):
    if pd.isna(text):
        return None
    text = str(text).lower()
    for category, pattern in compiled_patterns.items():
        if pattern.search(text):
            return category
    return None  # If no match
#rule for type classification
type_rules = {
    'CM':['Front', 'CCD back', 'TWP', 'Guide tire warning','steering','Malfunction'],
    'FC':['ground', 'power', 'inspection', 'checklist'],
    'MOD':['add','fco'],
    'P-CM':['IW','worn out', 'axle'],
    'PREP':['standby','headset','shunting','borrow']
}
 # pre-compile type patterns for speed
compiled_type_patterns = {}
for type_category, type_keywords in type_rules.items():
    # Use alternation (|) instead of positive lookahead for OR matching
    # Add word boundaries \b to match whole words
    escaped_keywords = [re.escape(keyword) for keyword in type_keywords]
    type_pattern = r'\b(' + '|'.join(escaped_keywords) + r')\b'
    compiled_type_patterns[type_category] = re.compile(type_pattern, flags=re.IGNORECASE)

def get_type_from_text(text):
    if pd.isna(text):
        return None
    text = str(text).lower()
    for type_category, type_pattern in compiled_type_patterns.items():
        if type_pattern.search(text):
            return type_category
    return None  # If no match

def process_excel(df):
    # Get today's date (as datetime)
    today = datetime.today()

    # Get start and end of the current week (Fri–Thurs)
    offset = (today.weekday() - 4) % 7  # 4 = Friday
    start_of_week = today - timedelta(days=offset)
    end_of_week = start_of_week + timedelta(days=6)

    # Filter using datetime comparisons (no .date() conversion)
    df = df[
        (df['Malfunction Start'].dt.date >= start_of_week.date()) &
        (df['Malfunction Start'].dt.date <= end_of_week.date())]
    # Processing logic (unchanged)
    s = df["Location"].astype(str).str
    last_two = s[-2:]
    is_sw = df["Description"].str.contains('SW', case=False, na=False)
    #not make table consider dash/- as formula
    df["Location"] = df["Location"].apply(
    lambda x: f"'{x}" if isinstance(x, str) and x.strip().startswith('-') else x)

    df["System"] = np.where(is_sw, "Track switch", "YM" + last_two)
    df["Week"] = df["Malfunction Start"].apply(date_to_week)
    df["Problem"] = df["Description"].apply(classify_text)
    df["Sub-system - Revised"]=""
    df["Type"]= df["Description"].apply(get_type_from_text)
    df["Root cause"]=""
    df["Corrective action"]=""
    df["Additional Description of Action"]=""
    # Keep original as datetime
    df["Malfunction Start"] = pd.to_datetime(df["Malfunction Start"])
    df["Malfunction End"] = pd.to_datetime(df["Malfunction End"])
    df["Sub-system - Functional location"] = df["Functional Location"].apply(get_code_text)
    # Column reordering
    cols = df.columns.tolist()
    cols.insert(5, cols.pop(cols.index("System")))
    cols.insert(6, cols.pop(cols.index("Sub-system - Functional location")))
    cols.insert(7, cols.pop(cols.index("Sub-system - Revised")))
    cols.insert(8, cols.pop(cols.index("Type")))
    cols.insert(9, cols.pop(cols.index("Problem")))
    cols.insert(10, cols.pop(cols.index("Root cause")))
    cols.insert(11, cols.pop(cols.index("Corrective action")))
    cols.insert(12, cols.pop(cols.index("Additional Description of Action")))
    cols.insert(17, cols.pop(cols.index("Week")))
    df = df[cols]
    return df

def main():
    st.set_page_config(page_title="Excel Processor", layout="wide")
    st.title("📊 Excel File Processor")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        try:
            # Read and process file
            df = pd.read_excel(uploaded_file)
            processed_df = process_excel(df)
            
            # Create formatted Excel output
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl', datetime_format="DD/MM/YYYY") as writer:
                processed_df.to_excel(writer, index=False, sheet_name="Processed Data")
                # Access the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets["Processed Data"]
                
                 # Define the table range
                last_row = worksheet.max_row
                last_col = worksheet.max_column
                end_col_letter = openpyxl.utils.get_column_letter(last_col)
                table_range = f"A1:{end_col_letter}{last_row}"
                
                # Create table with proper syntax
                tab = openpyxl.worksheet.table.Table(displayName="MyNewTable", ref=table_range)
                
                # Add style to the table
                style = openpyxl.worksheet.table.TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                
                # Add the table to the worksheet
                worksheet.add_table(tab)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Download button
            st.download_button(
                label="⬇️ Download Processed File",
                data=output.getvalue(),
                file_name=f"processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.error("Please ensure your file contains the required columns")

if __name__ == "__main__":
    main()